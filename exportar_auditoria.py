#!/usr/bin/env python3
"""
Exporta para EXCEL (.xlsx) todos os processos por trás da jurimetria, para AUDITORIA humana.
Abas: Resumo · Validados (êxito%) · Censo Juizado (amostras) · Ranking por réu · 1º grau DJEN.
Cada linha tem o nº do processo para conferência na fonte pública.
Saída: auditoria_dominia.xlsx
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def L(f,d=None): return json.load(open(f,encoding="utf-8")) if os.path.exists(f) else d

HEAD=Font(bold=True,color="FFFFFF"); FILL=PatternFill("solid",fgColor="7C1FD0")
def sheet(wb,nome,cols,linhas):
    ws=wb.create_sheet(nome)
    ws.append(cols)
    for c in range(1,len(cols)+1):
        cell=ws.cell(1,c); cell.font=HEAD; cell.fill=FILL; cell.alignment=Alignment(vertical="center")
    for row in linhas: ws.append(row)
    ws.freeze_panes="A2"
    for c in range(1,len(cols)+1):
        w=max(len(str(cols[c-1])), *(len(str(r[c-1])) for r in linhas[:400]) if linhas else [0])
        ws.column_dimensions[get_column_letter(c)].width=min(max(w+2,10),52)
    return ws

def main():
    wb=Workbook(); wb.remove(wb.active)
    RES={"219":"Procedência","220":"Improcedência","221":"Procedência em parte"}
    total=0

    # 1) Validados por IA (a base do % de êxito)
    RV=L("registros_validados.json",[]) or []
    linhas=[[r.get("proc"),r.get("setor"),r.get("tese"),r.get("ano"),r.get("reu"),r.get("foro"),
             r.get("resultado"),"Sim" if r.get("fav") else "Não","Sim" if r.get("dm") else "Não",r.get("valor")] for r in RV]
    sheet(wb,"Validados (exito%)",
          ["Nº do processo","Setor","Tese","Ano","Réu","Foro","Resultado","Consumidor venceu","Teve dano moral","Valor (R$)"],linhas)
    total+=len(linhas)

    # 2) Censo Juizado (Datajud) — amostras de processo por tese
    cj=L("censo_juizado.json",{}) or {}
    lin=[]
    for seg,dd in (cj.get("segmentos") or {}).items():
        rot=dd.get("rotulo",seg)
        for t in dd.get("por_tese",[]):
            for p in (t.get("ex_favoravel") or []):
                lin.append([p,rot,t["nome"],"Favorável (procedência/parcial)",t.get("pct_fav"),t.get("n")])
            for p in (t.get("ex_improcedente") or []):
                lin.append([p,rot,t["nome"],"Improcedente",t.get("pct_fav"),t.get("n")])
    sheet(wb,"Censo Juizado (amostras)",
          ["Nº do processo","Segmento","Tese/Assunto","Sentido da decisão","% favorável da tese","n da tese"],lin)
    total+=len(lin); n_censo=len(lin)

    # 3) Ranking por réu — exemplos
    rs=L("reus_stats.json",{}) or {}
    lin=[]
    for r in rs.get("reus",[]):
        for p in (r.get("ex_favoravel") or []):
            lin.append([p,r["reu"],r.get("setor"),"Favorável ao consumidor",r.get("pct_fav_sinal"),r.get("dano_moral_mediana"),r.get("n")])
        for p in (r.get("ex_desfavoravel") or []):
            lin.append([p,r["reu"],r.get("setor"),"Desfavorável",r.get("pct_fav_sinal"),r.get("dano_moral_mediana"),r.get("n")])
    sheet(wb,"Ranking por reu",
          ["Nº do processo","Réu","Setor","Sinal","% favorável do réu","Dano moral mediano","n do réu"],lin)
    total+=len(lin); n_reu=len(lin)

    # 4) 1º grau DJEN (inteiro teor) — base do dano moral/comarca
    dje=L("registros_dje.json",[]) or []
    lin=[[d.get("proc"),d.get("s"),d.get("tese"),d.get("res"),d.get("v"),d.get("c"),d.get("j"),d.get("r"),d.get("fg"),d.get("em"),d.get("link")] for d in dje]
    sheet(wb,"1o grau DJEN",
          ["Nº do processo","Setor","Tese","Resultado","Valor dano moral (R$)","Comarca/Órgão","Juiz","Réu","Foro","Ementa (resumo)","Link íntegra"],lin)
    total+=len(lin)

    # Resumo (primeira aba)
    ws=wb.create_sheet("Resumo",0)
    ws.append(["DOMINIA — Planilha de auditoria dos processos usados na jurimetria"])
    ws["A1"].font=Font(bold=True,size=14,color="5B16A8")
    ws.append([])
    ws.append(["Aba","O que contém","Nº de processos"])
    for c in range(1,4): ws.cell(3,c).font=Font(bold=True)
    ws.append(["Validados (exito%)","Decisões validadas que formam o % de êxito ao consumidor",len(RV)])
    ws.append(["Censo Juizado (amostras)","Amostras de nº de processo por tese (censo Datajud/CNJ)",n_censo])
    ws.append(["Ranking por reu","Exemplos de processos por réu (ranking de êxito)",n_reu])
    ws.append(["1o grau DJEN","Sentenças de 1º grau com inteiro teor (dano moral, comarca, juiz)",len(dje)])
    ws.append(["TOTAL","Linhas de processo para conferência",total])
    ws.append([])
    ws.append(["Cada nº de processo é público — confira na Consulta Processual do TJBA / PJe."])
    ws.append(["Metodologia: procedência por código oficial de movimento (219/221 x 220). Precedente é persuasivo, não vinculante."])
    ws.column_dimensions["A"].width=26; ws.column_dimensions["B"].width=64; ws.column_dimensions["C"].width=16

    wb.save("auditoria_dominia.xlsx")
    print(f"OK auditoria_dominia.xlsx — {total} linhas de processo no total")
    print(f"  Validados: {len(RV)} · Censo amostras + Réu + 1º grau DJEN: {len(dje)} sentenças com íntegra")

if __name__=="__main__": main()
